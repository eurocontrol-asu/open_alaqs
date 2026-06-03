"""Post-inventory validation: compare movements emissions CSV against the
validation reference xlsx for a chosen calculation method.

Run AFTER running the inventory in QGIS and exporting the per-movement
emissions to CSV via the Table View output module. CSV columns:

    timestamp, source_type, source_name, co_kg, co2_kg, hc_kg, nox_kg,
    sox_kg, pm10_kg, p1_kg, p2_kg, pm10_organic_kg, pm10_nonvol_kg,
    pm10_sul_kg, wkt

Per-movement totals are computed by aggregating all CSV rows for the
same ``source_name``. The tool then prints a side-by-side comparison
against the reference column for the chosen method, with delta % flagged
when outside 0.5%.

Usage:
    python tools/compare_inventory_to_reference.py <movements.csv>
        [--method=METHOD] [<reference.xlsx>]

Methods (must match the calculation method used in the plugin run):
    bymode        bymode reference column (default)
    bffm2_traj    BFFM2 trajectory reference column
    bffm2_anchor  BFFM2 anchor reference column

Default reference: example/training/training_validation_reference.xlsx

Exit code 0 if all movements match within 0.5%, non-zero otherwise.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
DEFAULT_REF = REPO / "example" / "training" / "training_validation_reference.xlsx"

# Column-offset within each 6-column pollutant block for each method.
# Reference sheet layout per pollutant: bymode, CAEP14 LTO, BFFM2 traj,
# CAEP14 BF(traj), BFFM2 anchor, CAEP14 BF(anch).
METHOD_OFFSET = {
    "bymode": 0,
    "bffm2_traj": 2,
    "bffm2_anchor": 4,
}


def load_movements_csv(csv_path: Path) -> dict[int, dict[str, float]]:
    """Aggregate per-movement emissions from the CSV."""
    by_oid: dict[int, dict[str, float]] = {}
    pollutants = ("co", "co2", "hc", "nox", "sox", "pm10")
    with csv_path.open() as f:
        for row in csv.DictReader(f):
            name = row.get("source_name", "")
            if not name.startswith("id "):
                continue
            try:
                oid = int(name.split(":", 1)[0].removeprefix("id ").strip())
            except ValueError:
                continue
            agg = by_oid.setdefault(oid, {p: 0.0 for p in pollutants})
            for p in pollutants:
                try:
                    agg[p] += float(row.get(f"{p}_kg", "0") or 0.0)
                except ValueError:
                    pass
    return by_oid


def load_reference(xlsx_path: Path, method: str) -> dict[int, dict[str, float]]:
    """Read per-movement reference emissions for the requested method."""
    import openpyxl

    if method not in METHOD_OFFSET:
        raise ValueError(f"Unknown method '{method}'. Valid: {list(METHOD_OFFSET)}")
    offset = METHOD_OFFSET[method]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    refs: dict[int, dict[str, float]] = {}

    ws = wb["Movement"]
    cols = {
        p: 3 + 6 * i + offset
        for i, p in enumerate(["co", "co2", "hc", "nox", "sox", "pm10"])
    }
    for r_idx in range(6, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
        mov_id = row[0]
        if not isinstance(mov_id, int):
            continue
        entry: dict[str, float] = {}
        for p, col in cols.items():
            val = row[col - 1]
            if isinstance(val, (int, float)):
                entry[p] = float(val)
        refs[mov_id] = entry

    # Helicopter FOCA Appendix A canonical (AS50, SINGLE_TURBOSHAFT, 732 SHP).
    # FOCA values are method-agnostic by construction.
    refs[14] = {
        "co": 0.20324,
        "co2": 41.94,
        "hc": 0.15886,
        "nox": 0.08456,
        "sox": 0.0,
        "pm10": 0.00264,
    }
    refs[15] = {
        "co": 0.13702,
        "co2": 37.75,
        "hc": 0.10879,
        "nox": 0.06530,
        "sox": 0.0,
        "pm10": 0.00209,
    }
    return refs


def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(1)

    method = "bymode"
    positional = []
    for a in args:
        if a.startswith("--method="):
            method = a.split("=", 1)[1].lower()
        else:
            positional.append(a)

    if not positional:
        print(__doc__)
        sys.exit(1)
    csv_path = Path(positional[0])
    ref_xlsx = Path(positional[1]) if len(positional) > 1 else DEFAULT_REF
    if not csv_path.exists():
        print(f"CSV file not found: {csv_path}", file=sys.stderr)
        sys.exit(1)
    if not ref_xlsx.exists():
        print(f"Reference file not found: {ref_xlsx}", file=sys.stderr)
        sys.exit(1)
    if method not in METHOD_OFFSET:
        print(
            f"Unknown method '{method}'. Valid: {list(METHOD_OFFSET)}",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Reference: {ref_xlsx}")
    print(f"Inventory: {csv_path}")
    print(f"Method:    {method}")
    print()

    plugin = load_movements_csv(csv_path)
    refs = load_reference(ref_xlsx, method)

    if not plugin:
        print("No movement rows found in CSV.")
        sys.exit(2)

    pollutants = ("co", "co2", "hc", "nox", "pm10")
    print(f"{'OID':>3}  ", end="")
    for p in pollutants:
        print(
            f"{p.upper() + ' ref':>11}  {p.upper() + ' plg':>11}  {'D%':>7}  ",
            end="",
        )
    print()
    print("-" * (5 + 35 * len(pollutants)))

    n_ok = n_warn = n_miss = 0
    for oid in sorted(refs):
        if oid not in plugin:
            n_miss += 1
            print(f"{oid:>3}  MOVEMENT MISSING FROM PLUGIN OUTPUT")
            continue
        print(f"{oid:>3}  ", end="")
        worst = 0.0
        for p in pollutants:
            r = refs[oid].get(p)
            v = plugin[oid].get(p)
            if r is None or v is None or r == 0:
                print(f"{'-':>11}  {'-':>11}  {'-':>7}  ", end="")
                continue
            pct = (v - r) / r * 100.0
            worst = max(worst, abs(pct))
            flag = "" if abs(pct) < 0.5 else ("!" if abs(pct) < 2.0 else "**")
            print(f"{r:>11.5f}  {v:>11.5f}  {pct:>+6.2f}{flag:<1} ", end="")
        print()
        if worst < 0.5:
            n_ok += 1
        else:
            n_warn += 1

    print()
    print(f"Movements within 0.5%:  {n_ok}")
    print(f"Movements outside 0.5%: {n_warn}")
    print(f"Movements missing:      {n_miss}")
    sys.exit(0 if (n_warn == 0 and n_miss == 0) else 1)


if __name__ == "__main__":
    main()
